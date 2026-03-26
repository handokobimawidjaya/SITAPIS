"""
Import script untuk data master dari file Excel.

Menimport data dari 3 file:
1. Data Master Jenis Naskah Dinas SIMONAS.xlsx → master_jenisnaskahdinas
2. Data Master Klasifikasi Arsip SIMONAS.xlsx → master_klasifikasiarsip
3. Data Master Unit Kerja SIMONAS.xlsx → master_unitkerja

Run dengan: python3 import_master_data.py
"""

import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'simonas.settings')
django.setup()

import openpyxl
from apps.master.models import JenisNaskahDinas, KlasifikasiArsip, UnitKerja


def import_jenis_naskah_dinas():
    """
    Import data dari 'Data Master Jenis Naskah Dinas SIMONAS.xlsx' ke JenisNaskahDinas.
    
    File berisi: Kode | Jenis Naskah Dinas | Status
    """
    filepath = 'document_requirement/Data Master Jenis Naskah Dinas SIMONAS.xlsx'
    
    print(f"\n📄 Import Jenis Naskah Dinas dari: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    created = 0
    updated = 0
    errors = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:  # Skip header
            continue

        kode = row[0] if len(row) > 0 else None
        nama = row[1] if len(row) > 1 else None
        status_str = row[2] if len(row) > 2 else 'Aktif'

        if not kode or not nama:
            continue

        try:
            is_active = str(status_str).strip().lower() == 'aktif'

            obj, created_flag = JenisNaskahDinas.objects.update_or_create(
                kode=str(kode).strip(),
                defaults={
                    'nama': str(nama).strip(),
                    'is_active': is_active,
                }
            )

            if created_flag:
                created += 1
            else:
                updated += 1

        except Exception as e:
            errors += 1
            print(f"  ✗ ERROR row {i} ({kode}): {e}")

    print(f"  ✓ Jenis Naskah Dinas: {created} dibuat, {updated} diperbarui, {errors} error")
    return created, updated, errors


def import_klasifikasi_arsip():
    """
    Import data dari 'Data Master Klasifikasi Arsip SIMONAS.xlsx' ke KlasifikasiArsip.
    
    File berisi: Kode Klasifikasi | Keterangan | Status
    Mendukung struktur hierarki (PP → PP.01 → PP.01.1)
    """
    filepath = 'document_requirement/Data Master Klasifikasi Arsip SIMONAS.xlsx'
    
    print(f"\n📄 Import Klasifikasi Arsip dari: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    created = 0
    updated = 0
    errors = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:  # Skip header
            continue

        kode, nama, status_str = row[0], row[1], row[2]

        if not kode or not nama:
            continue

        try:
            # Determine parent for hierarchical structure
            parent = None
            if '.' in str(kode):
                parts = str(kode).split('.')
                if len(parts) == 2:
                    parent_code = parts[0]
                else:
                    parent_code = '.'.join(parts[:-1])

                try:
                    parent = KlasifikasiArsip.objects.get(kode=parent_code)
                except KlasifikasiArsip.DoesNotExist:
                    parent = None

            # Determine jenis based on code pattern
            jenis = 'substantif'
            code_prefix = str(kode).split('.')[0]
            if code_prefix in ['UM', 'KEU', 'BMN', 'PEG', 'HUM', 'HUK', 'REN', 'WAS']:
                jenis = 'fasilitatif'

            is_active = str(status_str).strip().lower() == 'aktif'

            obj, created_flag = KlasifikasiArsip.objects.update_or_create(
                kode=str(kode).strip(),
                defaults={
                    'nama': str(nama).strip(),
                    'jenis': jenis,
                    'parent': parent,
                    'is_active': is_active,
                    'deskripsi': '',
                }
            )

            if created_flag:
                created += 1
            else:
                updated += 1

        except Exception as e:
            errors += 1
            print(f"  ✗ ERROR row {i} ({kode}): {e}")

    # Second pass: Update parent relationships
    print("  Updating parent relationships...")
    for item in KlasifikasiArsip.objects.all():
        if '.' in str(item.kode):
            parts = str(item.kode).split('.')
            if len(parts) == 2:
                parent_code = parts[0]
            else:
                parent_code = '.'.join(parts[:-1])
            
            try:
                parent = KlasifikasiArsip.objects.get(kode=parent_code)
                if item.parent_id != parent.id:
                    item.parent = parent
                    item.save(update_fields=['parent'])
            except KlasifikasiArsip.DoesNotExist:
                pass

    print(f"  ✓ Klasifikasi Arsip: {created} dibuat, {updated} diperbarui, {errors} error")
    return created, updated, errors


def import_unit_kerja():
    """
    Import data dari 'Data Master Unit Kerja SIMONAS.xlsx' ke UnitKerja.
    
    File berisi: Kode Satker | Nama Satker | Status
    Digunakan untuk SATKER user.
    """
    filepath = 'document_requirement/Data Master Unit Kerja SIMONAS.xlsx'
    
    print(f"\n📄 Import Unit Kerja dari: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    created = 0
    updated = 0
    errors = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:  # Skip header
            continue

        kode, nama, status_str = row[0], row[1], row[2]

        if not kode or not nama:
            continue

        try:
            # Convert kode to string (handle integer codes like 18, 1801)
            kode_str = str(int(kode)) if isinstance(kode, (int, float)) else str(kode)
            is_active = str(status_str).strip().lower() == 'aktif'

            obj, created_flag = UnitKerja.objects.update_or_create(
                kode=kode_str,
                defaults={
                    'nama': str(nama).strip(),
                    'is_active': is_active,
                }
            )

            if created_flag:
                created += 1
            else:
                updated += 1

        except Exception as e:
            errors += 1
            print(f"  ✗ ERROR row {i} ({kode}): {e}")

    print(f"  ✓ Unit Kerja: {created} dibuat, {updated} diperbarui, {errors} error")
    return created, updated, errors


def print_summary():
    """Print summary of all imported data."""
    from apps.master.models import JenisNaskahDinas, KlasifikasiArsip, UnitKerja
    
    print("\n" + "="*60)
    print("📊 RINGKASAN DATA DI DATABASE")
    print("="*60)
    print(f"  Jenis Naskah Dinas:  {JenisNaskahDinas.objects.count():>6} records")
    print(f"  Klasifikasi Arsip:   {KlasifikasiArsip.objects.count():>6} records")
    print(f"  Unit Kerja:          {UnitKerja.objects.count():>6} records")
    print("="*60)
    
    # Show sample data
    print("\n📋 Sample Jenis Naskah Dinas:")
    for j in JenisNaskahDinas.objects.filter(is_active=True).order_by('kode')[:5]:
        print(f"    {j.kode} - {j.nama}")
    
    print("\n📋 Sample Klasifikasi Arsip:")
    for k in KlasifikasiArsip.objects.filter(is_active=True).order_by('kode')[:5]:
        print(f"    {k.kode} - {k.nama}")
    
    print("\n📋 Sample Unit Kerja:")
    for u in UnitKerja.objects.filter(is_active=True).order_by('kode')[:5]:
        print(f"    {u.kode} - {u.nama}")
    
    print("\n✓ Import selesai!")


if __name__ == '__main__':
    print("="*60)
    print("🚀 MEMULAI IMPORT DATA MASTER")
    print("="*60)

    total_created = 0
    total_updated = 0
    total_errors = 0

    # Import Jenis Naskah Dinas
    created, updated, errors = import_jenis_naskah_dinas()
    total_created += created
    total_updated += updated
    total_errors += errors

    # Import Klasifikasi Arsip
    created, updated, errors = import_klasifikasi_arsip()
    total_created += created
    total_updated += updated
    total_errors += errors

    # Import Unit Kerja
    created, updated, errors = import_unit_kerja()
    total_created += created
    total_updated += updated
    total_errors += errors

    # Print summary
    print(f"\n📈 TOTAL: {total_created} dibuat, {total_updated} diperbarui, {total_errors} error")
    
    print_summary()
